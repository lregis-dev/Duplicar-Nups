# process.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def destacar_nups(file_bytes):
    """Processa um arquivo Excel em bytes e retorna um novo com NUPs duplicados destacados."""

    # Carrega workbook a partir de bytes
    wb = load_workbook(filename=BytesIO(file_bytes))
    ws = wb.active

    # Cores
    fill1 = PatternFill(start_color="FCA48C", end_color="FCA48C", fill_type="solid")  # Vermelho
    fill2 = PatternFill(start_color="A4DDFA", end_color="A4DDFA", fill_type="solid")  # Azul

    nup_colors = {}
    color_index = 0  
    nup_col_index = 1  # Coluna do NUP (segunda coluna da planilha)

    # Lista de NUPs
    all_nups = [
        ws.cell(row=i, column=nup_col_index + 1).value
        for i in range(2, ws.max_row + 1)
    ]

    # Contagem
    nup_counts = pd.Series(all_nups).value_counts()

    # Destacar duplicados
    for row_index in range(2, ws.max_row + 1):
        nup_cell = ws.cell(row=row_index, column=nup_col_index + 1)
        nup_value = nup_cell.value

        if pd.notna(nup_value) and nup_counts.get(nup_value, 0) > 1:
            if nup_value not in nup_colors:
                nup_colors[nup_value] = fill1 if color_index == 0 else fill2
                color_index = 1 - color_index

            nup_cell.fill = nup_colors[nup_value]

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
